import json
import csv
import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from backend.core.logging import logger

class ExportService:
    """Export extracted data in multiple formats"""

    def __init__(self):
        pass

    async def to_json(self, invoice: dict) -> dict:
        """Export single invoice as JSON"""
        return {
            "invoice_id": invoice["invoice_id"],
            "filename": invoice["filename"],
            "extracted_data": invoice.get("extracted_data", {}),
            "confidence_scores": invoice.get("confidence_scores", {}),
            "processing_status": invoice["processing_status"],
            "reviewed_at": invoice.get("reviewed_at"),
            "review_status": invoice.get("review_status")
        }

    async def batch_to_json(self, invoices: List[dict]) -> List[dict]:
        """Export multiple invoices as JSON array"""
        return [await self.to_json(inv) for inv in invoices]

    async def to_csv(self, invoices: List[dict]) -> str:
        """Export invoices as CSV"""
        output = io.StringIO()
        
        if not invoices:
            return ""
        
        # Get all unique field names
        all_fields = set()
        for invoice in invoices:
            all_fields.update(invoice.get("extracted_data", {}).keys())
            
        fieldnames = ["invoice_id", "filename", "status"] + sorted(all_fields)
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for invoice in invoices:
            extracted = invoice.get("extracted_data", {})
            row = {
                "invoice_id": invoice["invoice_id"],
                "filename": invoice["filename"],
                "status": invoice["processing_status"]
            }
            row.update(extracted)
            writer.writerow(row)
            
        return output.getvalue()

    async def to_excel(self, invoices: List[dict]) -> io.BytesIO:
        """Export invoices as Excel workbook"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        
        # Get all unique fields
        all_fields = set()
        for invoice in invoices:
            all_fields.update(invoice.get("extracted_data", {}).keys())
            
        headers = ["Invoice ID", "Filename", "Status"] + sorted(all_fields)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        # Write data
        for row_idx, invoice in enumerate(invoices, 2):
            extracted = invoice.get("extracted_data", {})
            
            ws.cell(row=row_idx, column=1, value=invoice["invoice_id"])
            ws.cell(row=row_idx, column=2, value=invoice["filename"])
            ws.cell(row=row_idx, column=3, value=invoice["processing_status"])
            
            for col_idx, field in enumerate(sorted(all_fields), 4):
                value = extracted.get(field, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
            
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
